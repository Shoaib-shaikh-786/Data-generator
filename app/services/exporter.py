import json
import csv
import io
import xml.etree.ElementTree as ET
from typing import Any, Dict, List, Optional
from app.models.schemas import ExportFormat


def flatten_dict(d: Dict[str, Any], parent_key: str = "", sep: str = ".") -> Dict[str, Any]:
    """Flatten nested dicts for CSV/Excel export."""
    items = {}
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.update(flatten_dict(v, new_key, sep=sep))
        else:
            items[new_key] = v
    return items


def export_to_json(records: List[Dict[str, Any]]) -> bytes:
    return json.dumps(records, indent=2, default=str).encode("utf-8")


def export_to_csv(records: List[Dict[str, Any]]) -> bytes:
    if not records:
        return b""
    flat = [flatten_dict(r) for r in records]
    all_keys = list(dict.fromkeys(k for row in flat for k in row.keys()))
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=all_keys, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(flat)
    return buf.getvalue().encode("utf-8")


def export_to_excel(records: List[Dict[str, Any]]) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

    if not records:
        wb = openpyxl.Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    flat = [flatten_dict(r) for r in records]
    all_keys = list(dict.fromkeys(k for row in flat for k in row.keys()))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")

    for col_idx, key in enumerate(all_keys, 1):
        cell = ws.cell(row=1, column=col_idx, value=key)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(key) + 4, 12)

    for row_idx, row in enumerate(flat, 2):
        for col_idx, key in enumerate(all_keys, 1):
            ws.cell(row=row_idx, column=col_idx, value=str(row.get(key, "")))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_to_xml(records: List[Dict[str, Any]], table_name: str = "table") -> bytes:
    root = ET.Element("dataset")
    root.set("table", table_name)
    root.set("count", str(len(records)))

    def dict_to_xml(parent: ET.Element, data: Any, tag: str = "field"):
        if isinstance(data, dict):
            for k, v in data.items():
                child = ET.SubElement(parent, str(k).replace(" ", "_"))
                dict_to_xml(child, v, tag)
        elif isinstance(data, list):
            for item in data:
                child = ET.SubElement(parent, "item")
                dict_to_xml(child, item, tag)
        else:
            parent.text = str(data) if data is not None else ""

    for record in records:
        record_el = ET.SubElement(root, "record")
        dict_to_xml(record_el, record)

    tree = ET.ElementTree(root)
    buf = io.BytesIO()
    ET.indent(tree, space="  ")
    tree.write(buf, encoding="utf-8", xml_declaration=True)
    return buf.getvalue()


def export_to_parquet(records: List[Dict[str, Any]]) -> bytes:
    try:
        import pandas as pd
    except ImportError:
        raise ImportError("pandas is required for Parquet export.")

    if not records:
        return b""

    flat = [flatten_dict(r) for r in records]
    df = pd.DataFrame(flat)
    buf = io.BytesIO()
    df.to_parquet(buf, index=False)
    return buf.getvalue()


def export_data(
    records: List[Dict[str, Any]],
    fmt: ExportFormat,
    table_name: str = "table",
) -> tuple[bytes, str, str]:
    """Returns (content_bytes, media_type, filename_extension)"""
    if fmt == ExportFormat.JSON:
        return export_to_json(records), "application/json", "json"
    elif fmt == ExportFormat.CSV:
        return export_to_csv(records), "text/csv", "csv"
    elif fmt == ExportFormat.EXCEL:
        return (
            export_to_excel(records),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "xlsx",
        )
    elif fmt == ExportFormat.XML:
        return export_to_xml(records, table_name), "application/xml", "xml"
    elif fmt == ExportFormat.PARQUET:
        return export_to_parquet(records), "application/octet-stream", "parquet"
    else:
        return export_to_json(records), "application/json", "json"